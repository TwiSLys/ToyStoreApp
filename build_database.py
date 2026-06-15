import os
import re
import sqlite3
from datetime import datetime

from openpyxl import load_workbook


ROOT = os.path.dirname(os.path.abspath(__file__))
DATABASE_DIR = os.path.join(ROOT, "database")
DATABASE_PATH = os.path.join(DATABASE_DIR, "toy_store.db")


def clean_date(value):
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    if text == "30.02.2025":
        return "28.02.2025"
    datetime.strptime(text, "%d.%m.%Y")
    return text


def lookup_id(connection, table_name, value):
    connection.execute(
        f"INSERT OR IGNORE INTO {table_name}(name) VALUES (?)", (str(value).strip(),)
    )
    return connection.execute(
        f"SELECT id FROM {table_name} WHERE name = ?", (str(value).strip(),)
    ).fetchone()[0]


def build_database():
    os.makedirs(DATABASE_DIR, exist_ok=True)
    if os.path.exists(DATABASE_PATH):
        os.remove(DATABASE_PATH)
    connection = sqlite3.connect(DATABASE_PATH)
    connection.execute("PRAGMA foreign_keys = ON")
    with open(os.path.join(DATABASE_DIR, "schema.sql"), encoding="utf-8") as schema:
        connection.executescript(schema.read())

    roles = ["Администратор", "Менеджер", "Авторизированный клиент"]
    connection.executemany("INSERT INTO roles(name) VALUES (?)", [(role,) for role in roles])

    users_book = load_workbook(
        os.path.join(DATABASE_DIR, "user_import.xlsx"), data_only=True
    ).active
    for role, full_name, login, password in list(users_book.values)[1:]:
        if not role:
            continue
        role_id = connection.execute(
            "SELECT id FROM roles WHERE name = ?", (role.strip(),)
        ).fetchone()[0]
        connection.execute(
            """
            INSERT INTO users(role_id, full_name, login, password)
            VALUES (?, ?, ?, ?)
            """,
            (role_id, full_name.strip(), login.strip(), password),
        )

    products_book = load_workbook(
        os.path.join(DATABASE_DIR, "Tovar.xlsx"), data_only=True
    ).active
    for row in list(products_book.values)[1:]:
        if not row[0]:
            continue
        article, name, unit, price, supplier, manufacturer, category = row[:7]
        discount, stock, description, photo = row[7:11]
        supplier_id = lookup_id(connection, "suppliers", supplier)
        manufacturer_id = lookup_id(connection, "manufacturers", manufacturer)
        category_id = lookup_id(connection, "categories", category)
        image_path = os.path.join("assets", "images", str(photo).upper())
        connection.execute(
            """
            INSERT INTO products(
                article, name, unit, price, supplier_id, manufacturer_id,
                category_id, discount, stock, description, image_path
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                article,
                name,
                unit,
                float(price),
                supplier_id,
                manufacturer_id,
                category_id,
                int(discount),
                int(stock),
                description,
                image_path,
            ),
        )

    points_book = load_workbook(
        os.path.join(DATABASE_DIR, "Пункты выдачи_import.xlsx"), data_only=True
    ).active
    for (address,) in points_book.values:
        if address:
            normalized = str(address).replace("\xa0", " ").replace("г.Лесной", "г. Лесной")
            connection.execute(
                "INSERT INTO pickup_points(address) VALUES (?)", (normalized,)
            )

    orders_book = load_workbook(
        os.path.join(DATABASE_DIR, "Заказ_import.xlsx"), data_only=True
    ).active
    for row in list(orders_book.values)[1:]:
        if not row[0]:
            continue
        number, item_text, order_date, delivery_date, point_number = row[:5]
        client_name, receive_code, status = row[5:8]
        client = connection.execute(
            """
            SELECT users.id FROM users
            JOIN roles ON roles.id = users.role_id
            WHERE users.full_name = ? AND roles.name = 'Авторизированный клиент'
            LIMIT 1
            """,
            (client_name,),
        ).fetchone()
        connection.execute(
            """
            INSERT INTO orders(
                id, order_date, delivery_date, pickup_point_id,
                client_id, receive_code, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(number),
                clean_date(order_date),
                clean_date(delivery_date),
                int(point_number),
                client[0] if client else None,
                int(receive_code),
                status.strip(),
            ),
        )
        parts = [part.strip() for part in str(item_text).split(",")]
        for index in range(0, len(parts), 2):
            article = parts[index]
            quantity = int(parts[index + 1])
            product_id = connection.execute(
                "SELECT id FROM products WHERE article = ?", (article,)
            ).fetchone()[0]
            connection.execute(
                """
                INSERT INTO order_items(order_id, product_id, quantity)
                VALUES (?, ?, ?)
                """,
                (int(number), product_id, quantity),
            )
    connection.commit()
    connection.close()
    print(f"Database created: {DATABASE_PATH}")


if __name__ == "__main__":
    build_database()
